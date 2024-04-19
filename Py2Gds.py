import openpyxl
import os
import json
import calculate
import sys
import gdspy
import time

# 功能：检查是否创建skill&gds文件夹，如果没有则创建文件夹
def init_dir():  
    output_dir_path = os.path.join(os.getcwd(), "output")    
  
    # 创建 skill 目录  
    if not os.path.exists(output_dir_path):  
        try:  
            os.mkdir(output_dir_path)    
        except Exception as exp:  
            print("创建output文件夹失败，请检查权限")
            return  
        
# 功能：根据输入表格信息加载模版
class Load_Template():
    def __init__(self, name: str, value: dict):
        current_dir = os.getcwd()
        templates_dir = os.path.join(current_dir, 'templates')
        full_path = os.path.join(templates_dir, name)
        # 检查文件是否存在  
        if os.path.isfile(full_path):  
            # 打开并读取文件  
            with open(full_path, 'r') as file:
                content = file.read()
            
            content_json = json.loads(content)
            self.parameter = content_json['parameter']
            self.expression = content_json['expression']
            self.circulate = content_json['circulate']
            self.coordinate = content_json['coordinate']
            self.condition = content_json['condition']
            self.value = value

            self.parameter_value = {}
            self.circulate_value = {}
            self.coordinate_value = []
            self.layermap = {}
            self.load_layermap()
            self.process_parameter()
            self.process_expression()
            self.process_condition()
            self.process_circulate()
            self.process_coordinate()

    # 功能：处理parameter项        
    def process_parameter(self):
        for item in self.parameter:
            if item in self.value:
                self.parameter_value[item] = self.value[item]
            else:
                print(f"Missing parameters: {item}")  
                sys.exit(2)
        return

    # 功能：处理expression项
    def process_expression(self):
        for item in self.expression:
            self.parameter_value[item] = calculate.count_total(self.expression[item], self.parameter_value)
        return

    # 功能：处理condition项
    def process_condition(self):
        for item in self.condition:
            if calculate.count_total(item, self.parameter_value):
                for t_operate in self.condition[item]['true']:
                    self.parameter_value[t_operate] = calculate.count_total(self.condition[item]['true'][t_operate], self.parameter_value)
            else:
                for f_operate in self.condition[item]['false']:
                    self.parameter_value[f_operate] = calculate.count_total(self.condition[item]['false'][f_operate], self.parameter_value)
        return
    
    # 功能：处理circulate项
    def process_circulate(self):
        for item in self.circulate:
            start = calculate.count_total(self.circulate[item][0], self.parameter_value)
            end = calculate.count_total(self.circulate[item][2], self.parameter_value)
            step = calculate.count_total(self.circulate[item][1], self.parameter_value)
            self.circulate_value[item] = [i for i in range(start, end, step)]
        return 
    
    # 功能：处理coordinate项
    def process_coordinate(self):
        temp = []
        for item in self.coordinate:
            temp.append([calculate.replace_variables(item[0], self.parameter_value), calculate.replace_variables(item[1], self.parameter_value), calculate.replace_variables(item[2], self.parameter_value), calculate.replace_variables(item[3], self.parameter_value), calculate.replace_variables(item[4], self.parameter_value), calculate.replace_variables(item[5], self.parameter_value)])

        skt = []
        for item in range(len(temp)):
            if temp[item][4] in self.layermap:
                if temp[item][5] in self.layermap[temp[item][4]]:
                    skt = self.layermap[temp[item][4]][temp[item][5]]
                    temp[item][4] = skt[0]
                    temp[item][5] = skt[1]
        step = []
        for string in temp:
            step.append(calculate.cir_total(string, self.circulate_value))

        for item in step:
            self.coordinate_value += item
        

            
        return
    
    # 功能：加载layermap文件
    def load_layermap(self):
        temp = []
        with open('/Users/zuowang/Py2Gds/layermap/hlmc22fs.txt', 'r') as file:
            lines = file.readlines()
            for line in lines:
                temp.append(line.strip().split())  
      
            for item in temp:  
                if item[0] in self.layermap:
                    self.layermap[item[0]][item[1]] = [item[2], item[3]]
                else:
                    self.layermap[item[0]] = {}
                    self.layermap[item[0]][item[1]] =  [item[2], item[3]]
  
            # 将defaultdict转换为普通字典（如果需要的话）
        return
    
# 功能：加载excel
def load_excel(path: str):
    if not os.path.exists(path):
        print(f"错误：文件{path}不存在")
        sys.exit(2)
    else:
        excel = openpyxl.load_workbook(path)
        all_sheet = excel.sheetnames
    data = []
    for item in all_sheet:
        sheet = excel[item]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):  
                sheet_data.append(row) 
        l = len(sheet_data)
        w = len(sheet_data[0])
        temp = []
        for i in range(l - 1):
            dict_v = {}
            for j in range(w):
                dict_v[sheet_data[0][j]] = sheet_data[i + 1][j]
            temp.append(dict_v)
        list = [item, temp]
        data.append(list)
    return data

# 功能：制作gds名字
def dict_to_string(dictionary):  
    return ''.join(f'_{key}{value}' for key, value in sorted(dictionary.items())) 

# 功能：将列表转换成元组
def list_to_tuple(origin: list):
    temp = []
    for item in origin:
        tuple1 = (item[0], item[1])
        tuple2 = (item[2], item[3])
        temp.append([tuple1, tuple2, item[4], item[5]])
    return temp

if __name__ == '__main__':
    if len(sys.argv) != 2:
       print("错误：请提供excel文件地址")
       sys.exit(1)

    start_time = time.time()
    init_dir()
    excel_path = sys.argv[1]
    data = load_excel(excel_path)
    print("导入表格成功")
    list_temp = {}
    for item in data:
        for i in item[1]:
            print(f"正在加载：{item[0]}_{dict_to_string(i)[:10]}")
            list_temp[item[0] + dict_to_string(i)] = Load_Template(item[0] + '.json',i)

    tuple_temp = {}
    for item in list_temp:
        tuple_temp[item] = list_to_tuple(list_temp[item].coordinate_value)
    lib = []
    cell = []
    i = 0
    for item in tuple_temp:
        print(f"正在绘制版图：{item[:15]}")
        lib.append(gdspy.GdsLibrary())
        cell.append(lib[i].new_cell(item))
        for data in tuple_temp[item]:
            cell[i].add(gdspy.Rectangle(data[0],data[1],data[2],data[3]))
        lib[i].write_gds('./output/' + item + '.gds')
        i = i + 1
    end_time = time.time()
    running_time = end_time - start_time  
    print("程序运行时间: %s 秒" % running_time)
